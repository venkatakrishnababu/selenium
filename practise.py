import time

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
import openpyxl
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.devtools.v133.dom import scroll_into_view_if_needed
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import requests as requests
# serv_obj=Service(r"D:\Downloads\chromedriver-win64\chromedriver-win64\chromedriver.exe")
# driver=webdriver.Chrome(service=serv_obj)
# file=r"D:\Downloads\DDS.xlsx"
# ##file-->workbook-->sheet-->rows/coloums
# workbook=openpyxl.load_workbook(file)
# sheetname=workbook['Sheet1']
# print(sheetname.max_row)
# print(sheetname.max_column)
# sheetname.cell(2,1).value='def'
# sheetname.cell(2,2).value='def'
# sheetname.cell(2,3).value='def'
# workbook.save(file)
# b=sheetname.cell(2,1).value
# print(b)


# try:
#     element=wait.until(EC.presence_of_element_located((By.LINK_TEXT,"Bank Project")))
#     #element=wait.until(EC.element_to_be_clickable((By.XPATH,"//a[normalize-space()='Bank Project']")))
#     #element = wait.until(EC.visibility_of_element_located((By.LINK_TEXT, "Bank Project")))
#     #element = wait.until(EC.visibility_of_element_located((By.LINK_TEXT, "Bank Project")))
#
#     print("element found")
# except:
#     print('element not  found')
# driver.find_element(By.LINK_TEXT,"Bank Project").click()
# time.sleep(5)

##download

# perfs={"download.default_directory":os.getcwd()}
# b=webdriver.ChromeOptions()
# b.add_experimental_option("perfs",perfs)

driver=webdriver.Chrome()
# driver.get("http://www.deadlinkcity.com/")
# b=driver.find_elements(By.XPATH,"//a")
# print(len(b))
# c=0
# d=0
# for i in b:
#     broken=i.get_attribute('href')
#     try:
#         res=requests.head(broken)
#     except:
#         None
#     if res.status_code>400:
#         c+=1
#         print("i am broken link:-",broken,res.status_code)
#     else:
#         d+=1
#         print("Iam valid link:-",broken)
# print(len(b))
# print(c,"iam broken link count")
# print(d,"Iam valid link count")
# driver.get("https://demo.guru99.com/test/simple_context_menu.html")
# element=driver.find_element(By.LINK_TEXT,"Bank Project").click()
# #driver.execute_script("arguments[0].scrollIntoView();",element)
# #element.send_keys(Keys.END)
# body = driver.find_element("tag name", "body")
# body.send_keys(Keys.END)  # Simulates pressing the "End" key
# time.sleep(10)
driver.implicitly_wait(10)
driver.get("https://www.naukri.com/companies-hiring-in-india?src=gnbCompanies_homepage_srch")
time.sleep(5)
# driver.find_element(By.XPATH,"//span[@class='ni-gnb-icn ni-gnb-icn-search']").click()
# driver.find_element(By.XPATH,"//input[@placeholder='Enter keyword / designation / companies']").send_keys('cloud')
# time.sleep(5)
# b=driver.find_elements(By.XPATH,"//*[@class='drop-layer suggestor-drop-layer  direction-bottom ']/div/ul/li")
# for i in b:
#     if 'Migration' in i.text:
#         i.click()
# driver.find_element(By.XPATH,"//*[@class='nI-gNb-sb__expDD']").click()
# c=driver.find_elements(By.XPATH,"//*[@id='sa-dd-scrollexperienceDD']/div/ul/li")
# for i in c:
#     if i.text=='4 years':
#         i.click()
#         break
# driver.find_element(By.XPATH,"//input[@placeholder='Enter location']").send_keys("pune,bengaluru")
# driver.find_element(By.XPATH,"//button[@class='nI-gNb-sb__icon-wrapper']").click()
# b=driver.find_element(By.XPATH,"//div[@class='styles_filterOptns__1vq77' and @data-filter-id='experience']")
# ops=ActionChains(driver)
# ops.drag_and_drop_by_offset(b,7,7)
# time.sleep(15)
# driver.find_element(By.XPATH,"//span[normalize-space()='+90 more']").click()
# b=driver.find_elements(By.XPATH,"//*[@class='swiper-wrapper']/ul/li/a/i")
# for i in b:
#     i.click()
# driver.find_element(By.XPATH,"//span[normalize-space()='+71 more']").click()
# time.sleep(15)
driver.find_element(By.XPATH,"//span[normalize-space()='+72 more']").click()
time.sleep(15)