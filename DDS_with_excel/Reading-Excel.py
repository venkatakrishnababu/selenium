#Links--> Workbook-->Sheets-->Rows--->Cells
### Reading the data from the excel sheet
from multiprocessing.pool import worker

import openpyxl
File=r'D:\\chrome\\testing_read.xlsx'
# workbook=openpyxl.load_workbook(Links)
# sheet=workbook["funny"]
# row=sheet.max_row
# col=sheet.max_column
# for i in range(1,row+1):
#     for j in range(1,col+1):
#         data=sheet.cell(i,j).value
#         print(data,end='    ')
#     print()

##########count all the rows
def rows(file,sheet):
    workbook=openpyxl.load_workbook(file)
    sheetname=workbook[sheet]
    return sheetname.max_row
print(rows(File,'funny'))

###Count all the coloums
def COL(file,sheet):
    workbook=openpyxl.load_workbook(file)
    sheetname=workbook[sheet]
    return sheetname.max_column
print(COL(File,'funny'))

### Read the data
def get_read(file,sheet,rows,coloums):
    workbook=openpyxl.load_workbook(file)
    sheetname=workbook[sheet]
    return sheetname.cell(rows,coloums).value
print(get_read(File,'funny',2,2))

###write the data
def get_write(file,sheet,rows,coloum,data):
    workbook=openpyxl.load_workbook(file)
    sheetname=workbook[sheet]
    sheetname.cell(rows,coloum).value = data
    workbook.save(file)
get_write(File,"funny",10,2,"goat")