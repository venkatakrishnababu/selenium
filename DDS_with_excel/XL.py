import openpyxl
class excel():
    ### Count the number of rows
    def get_rows(file,sheet):
        workbook = openpyxl.load_workbook(file)
        sheetname = workbook[sheet]
        return sheetname.max_row

    ### count the number of the coloums
    def get_col(file, sheet):
        workbook = openpyxl.load_workbook(file)
        sheetname = workbook[sheet]
        return sheetname.max_column

    #### read the data from the excel sheet
    def read_data(file, sheet, rows, coloums):
        workbook = openpyxl.load_workbook(file)
        sheetname = workbook[sheet]
        return sheetname.cell(rows, coloums).value

    ### write the data to the excel sheet
    def write_data(file, sheet, rows, coloums, data):
        workbook = openpyxl.load_workbook(file)
        sheetname = workbook[sheet]
        sheetname.cell(rows, coloums).value = data
        workbook.save(file)