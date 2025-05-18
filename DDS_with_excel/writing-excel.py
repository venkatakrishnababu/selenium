#Links--> Workbook-->Sheets-->Rows--->Cells
import openpyxl
#### TO write same data into the excel in the rows and coloums

# file=r"D:\\chrome\\testing.xlsx"
# workbook=openpyxl.load_workbook(file)
# sheet=workbook.active               #### sheet=workbook["Data"]
# for R in range(1,7):
#     for C in range(1,4):
#         sheet.cell(R,C).value='welcome'
# workbook.save(file)

#### TO write same data into the excel in the rows and coloums
file=r'D:\\chrome\\testing_read.xlsx'
workbook=openpyxl.load_workbook(file)
sheet=workbook.active
sheet.cell(7,1).value='106'
sheet.cell(7,2).value='fox'
sheet.cell(7,3).value='23'

sheet.cell(8,1).value='107'
sheet.cell(8,2).value='giraffe'
sheet.cell(8,3).value='29'

sheet.cell(9,1).value='108'
sheet.cell(9,2).value='Hippo'
sheet.cell(9,3).value='112'
workbook.save(file)

